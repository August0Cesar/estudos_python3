from openpyxl import load_workbook
import json


def create_lista_keys(file):
    wb = load_workbook('Arquivos_2.xlsx') # abrindo o Workbook test.xlsx
    ws = wb['Sheet']
    
    lista_de_keys = []
    for line in ws: 
        key = line[6].value.strip()
        if key not in lista_de_keys:
            lista_de_keys.append(line[6].value.strip())
    
    
    lista_de_keys.pop(0)
    return lista_de_keys


def retorna_dicionario(lista_valores):
    id = 1
    dicionario = {}
    for perguntas in lista_valores:
        dicionario[id] = perguntas
        id += 1
    return dicionario


def create_dicionario(file, lista_de_keys):
    dic = {}
    
    
    for key in lista_de_keys:
        
        lista_valores = []
        cont = 0
        

        wb = load_workbook('Arquivos_2.xlsx') # abrindo o Workbook test.xlsx
        ws = wb['Sheet']
        for line in ws:
            if cont != 0:
                line_key = line[6].value.strip()
                line_value = line[7].value.strip() if line[7].value != None and line[7].value != '' else 'Valor em branco'
                if line_key == key and line_value not in lista_valores and line_value != '':
                    lista_valores.append(line_value)
            cont += 1

        
        dic[key] = retorna_dicionario(lista_valores)
        # break
    
    # print(dic)
    print(json.dumps(dic,ensure_ascii=False, sort_keys=True, indent=4))




lista_keys = create_lista_keys('codigo_pergunta.csv')
create_dicionario('codigo_pergunta.csv',lista_keys)
print(lista_keys)


